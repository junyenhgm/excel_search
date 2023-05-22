import sys
import openpyxl

def search_keyword_in_excel(file_path, keyword):
    workbook = openpyxl.load_workbook(file_path)
    matched_sheets = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and keyword in str(cell.value):
                    matched_sheets.append(sheet_name)
                    break

    return matched_sheets

# Check for input
argc = len(sys.argv)
if argc != 3:
    print("excel_search.py [excel_name] [keyword]")
    exit()

# Get excel file and keyword from arguments
file_path = sys.argv[1]
keyword = sys.argv[2]

# Example usage
# file_path = 'sample.xlsx'
# keyword = 'Saudi Arabia'
matched_sheets = search_keyword_in_excel(file_path, keyword)
print(f"Keyword '{keyword}' is found in the following sheets: {matched_sheets}")
